import os
import xml.etree.ElementTree as ET
import tkinter.filedialog
import csv
import pandas as pd
import openpyxl
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import Font

ET.register_namespace('', 'http://soap.sforce.com/2006/04/metadata')
nsp = '{http://soap.sforce.com/2006/04/metadata}'

fieldToPermissionsForOutput = {'Headers':['Label','Type','Description','Inline Help Text']}
objectToPermissionsForOutput = {'Headers':[]}
userPermissionsForOutput = {'Headers':[]}
visualforcePagePermissionsForOutput = {'Headers':[]}
apexClassPermissionsForOutput = {'Headers':[]}
objectFieldDetailMap = {} # {object: {field: [label, type, description]}}
permSubFolders = ['/profiles','/permissionsets',]


def read_object_file(file_path):
    objectName = file_path.rsplit('/', 1)[-1][:-7]
    objectFieldDetailMap[objectName] = {}
    tree = ET.parse(file_path)
    root = tree.getroot()
    for elem in root.findall(nsp+'fields'):
        fieldName = elem.find(nsp+'fullName').text
        if elem.find(nsp+'label') is not None:
            objectFieldDetailMap[objectName][fieldName] = [elem.find(nsp+'label').text]
        else:
            objectFieldDetailMap[objectName][fieldName] = ['-']
        add_additional_field_information(elem, objectName, fieldName, 'type')
        add_additional_field_information(elem, objectName, fieldName, 'description')
        add_additional_field_information(elem, objectName, fieldName, 'inlineHelpText')


def add_additional_field_information(elem, objectName, fieldName, elemText):
    if elem.find(nsp+elemText) is not None:
        objectFieldDetailMap[objectName][fieldName].append(elem.find(nsp+elemText).text)
    else:
        objectFieldDetailMap[objectName][fieldName].append('-')


def read_permission_file(file_path, file_name):
    tree = ET.parse(file_path)
    root = tree.getroot()

    fieldKeys = set(fieldToPermissionsForOutput.keys())
    fieldKeys.discard('Headers')
    fieldToPermissionsForOutput['Headers'].append(file_name)
    for elem in root.findall(nsp+'fieldPermissions'):
        elem_text = elem.find(nsp+'field').text
        fieldKeys.discard(elem_text)
        if elem_text not in fieldToPermissionsForOutput:
            objFieldList = elem_text.rsplit('.')
            try:
                fieldData = objectFieldDetailMap[objFieldList[0]][objFieldList[1]]
                fieldToPermissionsForOutput[elem_text] = [fieldData[0],fieldData[1],fieldData[2],fieldData[3]]
            except KeyError as e:
                print(e)
                print('Skipped over the key since data was not found.')
                fieldToPermissionsForOutput[elem_text] = ['-','-','-','-']
            if (len(fieldToPermissionsForOutput['Headers']) > 5):
                counter = 5
                while counter < len(fieldToPermissionsForOutput['Headers']):
                    fieldToPermissionsForOutput[elem_text].append('-')
                    counter += 1

        if elem.find(nsp+'editable').text == 'true':
            fieldToPermissionsForOutput[elem_text].append('Edit')
        elif elem.find(nsp+'readable').text == 'true':
            fieldToPermissionsForOutput[elem_text].append('Read')
        else:
            fieldToPermissionsForOutput[elem_text].append('-')
    for elem in fieldKeys:
        fieldToPermissionsForOutput[elem].append('-')

    objectKeys = set(objectToPermissionsForOutput.keys())
    objectKeys.discard('Headers')
    objectToPermissionsForOutput['Headers'].append(file_name)
    for elem in root.findall(nsp+'objectPermissions'):
        elem_text = elem.find(nsp+'object').text
        objectKeys.discard(elem_text)
        if elem_text not in objectToPermissionsForOutput:
            objectToPermissionsForOutput[elem_text] = []
            if (len(objectToPermissionsForOutput['Headers']) > 1):
                counter = 1
                while counter < len(objectToPermissionsForOutput['Headers']):
                    objectToPermissionsForOutput[elem_text].append('-')
                    counter += 1

        if elem.find(nsp+'modifyAllRecords').text == 'true':
            objectToPermissionsForOutput[elem_text].append('ModifyAll')
        else:
            access_string = ''
            if elem.find(nsp+'allowCreate').text == 'true':
                access_string = access_string +'C '
            if elem.find(nsp+'allowRead').text == 'true':
                access_string = access_string +'R '
            if elem.find(nsp+'allowEdit').text == 'true':
                access_string = access_string +'U '
            if elem.find(nsp+'allowDelete').text == 'true':
                access_string = access_string +'D '
            if elem.find(nsp+'viewAllRecords').text == 'true':
                access_string = access_string +'VA'
            if len(access_string) > 0:
                objectToPermissionsForOutput[elem_text].append(access_string.strip())
            else:
                objectToPermissionsForOutput[elem_text].append('-')
    for elem in objectKeys:
        objectToPermissionsForOutput[elem].append('-')

    analyze_permissions(root, userPermissionsForOutput, 'userPermissions', 'name')
    analyze_permissions(root, visualforcePagePermissionsForOutput, 'pageAccesses', 'apexPage')
    analyze_permissions(root, apexClassPermissionsForOutput, 'classAccesses', 'apexClass')


def analyze_permissions(treeRoot, permissionMap, nodeType, nameNode):
    permKeys = set(permissionMap.keys())
    permKeys.discard('Headers')
    permissionMap['Headers'].append(file_name)
    for elem in treeRoot.findall(nsp+nodeType):
        elem_text = elem.find(nsp+nameNode).text
        permKeys.discard(elem_text)
        if elem_text not in permissionMap:
            permissionMap[elem_text] = []
            if (len(permissionMap['Headers']) > 1):
                counter = 1
                while counter < len(permissionMap['Headers']):
                    permissionMap[elem_text].append('-')
                    counter += 1

        if elem.find(nsp+'enabled').text == 'true':
            permissionMap[elem_text].append('True')
        else:
            permissionMap[elem_text].append('-')
    for elem in permKeys:
        permissionMap[elem].append('-')


def write_output_permission_file():
    macroExists = os.path.isfile('MacroTemplate.xlsm')
    wb = openpyxl.Workbook()
    if macroExists:
        wb = openpyxl.load_workbook(filename = 'MacroTemplate.xlsm', keep_vba = True)
    wsf = wb.active
    wsf.title = 'Field Permissions'
    wso = wb.create_sheet('Object Permissions')
    wsu = wb.create_sheet('User Permissions')
    wsc = wb.create_sheet('Class Permissions')
    wsp = wb.create_sheet('Page Permissions')
    populate_format_worksheet(wsf, fieldToPermissionsForOutput)
    populate_format_worksheet(wso, objectToPermissionsForOutput)
    populate_format_worksheet(wsu, userPermissionsForOutput)
    populate_format_worksheet(wsc, apexClassPermissionsForOutput)
    populate_format_worksheet(wsp, visualforcePagePermissionsForOutput)
    try:
        if macroExists:
            wb.save('DataDictionaryResults.xlsm')
        else:
            wb.save('DataDictionaryResults.xlsx')
    except PermissionError as e:
        print(e)
        print('Please make sure to close out of the DataDictionaryResults.xlsx file before running this tool.')


def populate_format_worksheet(worksheet, dataInput):
    df = pd.DataFrame.from_dict(data=dataInput, orient='index')
    for r in dataframe_to_rows(df, index=True, header=False):
        worksheet.append(r)
    worksheet.delete_rows(1,1)
    worksheet['A1'].value = 'API Name'
    header_font = openpyxl.styles.Font(bold=True)
    for cell in worksheet['1:1']:
        cell.font = header_font
    for cell in worksheet['A:A']:
        cell.font = header_font
    cell = worksheet['B2']
    worksheet.freeze_panes = cell


def set_base_folder():
    folder_path = tkinter.filedialog.askdirectory(title = 'Select the src or main folder')
    while folder_path != '' and not folder_path.lower().endswith('src') and not folder_path.lower().endswith('main'):
        print('Please select the src folder if it is in Metadata API format or the main folder if it is in DX format')
        folder_path = tkinter.filedialog.askdirectory(title = 'Select the src or main folder')
    if folder_path.endswith('main'):
        folder_path += '/default'
    return folder_path


# Begin execution
tkinter.Tk().withdraw()
base_folder_path = set_base_folder()
    
for file_name in os.listdir(base_folder_path+'/objects'):
    if not file_name.endswith('.resource-meta.xml'):
        read_object_file(base_folder_path+'/objects/'+file_name)

for folder in permSubFolders:
    for file_name in os.listdir(base_folder_path+folder):
        if not file_name.endswith('.resource-meta.xml'):
            read_permission_file(base_folder_path+folder+'/'+file_name, file_name)
write_output_permission_file()
